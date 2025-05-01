# utils.py

import openpyxl
from pathlib import Path
from typing import List, Dict

def load_script_list(
    path: Path,
    sheet_name: str = "Sheet1",
    start_row: int = 2,
    no_col: str = "A",      # ← 追加
    text_col: str = "B"     # ← 追加
) -> List[Dict[str, str]]:
    """
    Excel から
      - 台詞番号列(no_col)
      - 台詞内容列(text_col)
    を読み取り、以下の形式のリストを返します。
      [ {'no': '001', 'text': 'よし、行くぞ！'}, … ]
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    scripts = []
    row = start_row
    while True:
        no = ws[f"{no_col}{row}"].value
        txt = ws[f"{text_col}{row}"].value
        if no is None and txt is None:
            break  # 終端
        if no and txt:
            scripts.append({
                "no": str(no).zfill(3),    # 例: 1 → "001"
                "text": str(txt).strip()
            })
        row += 1

    return scripts
